'''
Author: sean seanzq0331@163.com
Date: 2024-06-25 12:08:07
LastEditors: sean seanzq0331@163.com
LastEditTime: 2024-06-26 15:55:04
FilePath: /queryScore/query.py
Description: 这是默认设置,请设置`customMade`, 打开koroFileHeader查看配置 进行设置: https://github.com/OBKoro1/koro1FileHeader/wiki/%E9%85%8D%E7%BD%AE
'''
#!/usr/bin/python
# -*- coding: UTF-8 -*-

import openpyxl.workbook
import requests
import datetime
import os
import ddddocr
from PIL import Image
import json
import openpyxl
from Students import Students
from time import sleep
from fake_useragent import UserAgent
import re


ocr = ddddocr.DdddOcr(show_ad=False)
session = requests.Session()
host = 'http://182.151.23.226:18004/gacjcx/'
headers = {'user-agent': UserAgent().chrome,
        'Host': '182.151.23.226:18004',
        'Referer': 'http://182.151.23.226:18004/gacjcx/'
}
host_res = session.get(host+'verifyCode.do?d'+str(int(datetime.datetime.now().timestamp()*1000)),headers=headers)
cookies = session.cookies.get_dict()
print(f'cookies:{cookies}')
queryInterface = 'query.do'
verifyInterface= 'verifyCode.do'
current_dir = os.path.dirname(os.path.abspath(__file__))
sheetPath = '中考成绩查询.xlsx'


def gen_code(img):
  with open(img, 'rb') as f:
    img_bytes = f.read()
  res = ocr.classification(img_bytes)
  hand = re.sub(r'[A-Za-z]',res,"")
  print(f'验证码:{res},handCode:{hand}')
  return hand[-4:]


def save_image(url, filename):
  """
  从给定的 URL 下载图片并保存到本地。

  Args:
      url (str): 图片 URL
      filename (str): 要保存的文件名
  """
  try:
    # 使用 requests 获取图片数据
    response = requests.get(url, stream=True,headers=headers,cookies=cookies)
    response.raise_for_status()  # 检查请求是否成功

    # 打开文件并保存图片数据
    with open(filename, 'wb') as f:
      for chunk in response.iter_content(chunk_size=8192): 
        if chunk:  # 过滤掉空 chunk
          f.write(chunk)
    print(f"图片已保存到 {filename}")

  except requests.exceptions.RequestException as e:
    print(f"下载图片失败: {e}")




workbook = openpyxl.load_workbook(sheetPath)
sheet = workbook.active
# 遍历每一行并输出 A 列的值
minrow=sheet.min_row #最小行
maxrow=sheet.max_row #最大行
mincol=sheet.min_column #最小列
maxcol=sheet.max_column #最大列

#按行读取
stu_list = []
for i in range(minrow,maxrow+1):
	if i>1:
		stu = Students()
		for j in range(mincol,maxcol+1):
					cell=str(sheet.cell(i,j).value)
					# print(cell,end=" ")
					if(j==1):
						stu.classNum=cell.strip()
					elif(j==2):
						stu.xm=cell.strip()
					elif(j==3):
						stu.bmh=cell.strip()
					elif(j==4):
						stu.zkzh=cell.strip()
					elif(j==5):
						stu.sfzh=cell.strip()
					else:
						continue
		stu_list.append(stu)
	else:
		continue


# 示例使用

for stu in stu_list:
	nowTime = str(int(datetime.datetime.now().timestamp()*1000))
	image_url = host+verifyInterface+"?d="+nowTime
	name = nowTime+".jpg"
	save_image(image_url,name)
	code = gen_code(name)
	os.remove(name)
	queryParams = {'bmh':stu.bmh,'zkzh':stu.zkzh,'sfzh':stu.sfzh,'verify':code}
	res = requests.post(host+queryInterface,data=queryParams,headers=headers,cookies=cookies)
	sleep(.3)
	jsonRes = json.loads(res.text)
	print(f'jsonRes:{jsonRes},type:{type(jsonRes)}')
	if(jsonRes['code']=="0"):
		jsonData = jsonRes['data']
		stu.yw=jsonData['yw']
		stu.sx=jsonData['sx']
		stu.yy=jsonData['yy']
		stu.wl=jsonData['wl']
		stu.hx=jsonData['hx']
		stu.dl=jsonData['dl']
		stu.ls=jsonData['ls']
		stu.sw=jsonData['sw']
		stu.ddyfz=jsonData['ddyfz']
		stu.fjf=jsonData['fjf']
		stu.tyyjk=jsonData['tyyjk']
		

sheetTitle = ['班级','姓名','报名号','准考证号','身份证号','语文','数学','英语','物理','化学','生物','道德与法治','历史','地理','体育与健康','加分','总分']
result = openpyxl.Workbook()
sheet_result = result.active
sheet_result.append(sheetTitle)
for stu in stu_list:
	print(f'dictstu:{dict}')
	sheet_result.append(dict(stu))

result.save('中考成绩查询结果.xlsx')