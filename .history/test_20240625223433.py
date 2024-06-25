'''
Author: sean seanzq0331@163.com
Date: 2024-06-25 12:08:07
LastEditors: sean seanzq0331@163.com
LastEditTime: 2024-06-25 22:34:33
FilePath: /queryScore/test.py
Description: 这是默认设置,请设置`customMade`, 打开koroFileHeader查看配置 进行设置: https://github.com/OBKoro1/koro1FileHeader/wiki/%E9%85%8D%E7%BD%AE
'''
import openpyxl
from Students import Students
res = {"code":"0","data":{"bmh":"246305102696","ddyfz":"55","dl":"37","fjf":"0","hx":"59","ls":"39","sfzh":"511623*********728","sw":"49","sx":"107","tyyjk":"64","wl":"55","xm":"包**","yw":"104","yy":"116","zf":"685","zkzh":"630508921"}}

path = '中考成绩查询.xlsx'

workbook = openpyxl.load_workbook(path)
sheet = workbook.active
# 遍历每一行并输出 A 列的值
minrow=sheet.min_row #最小行
maxrow=sheet.max_row #最大行
mincol=sheet.min_column #最小列
maxcol=sheet.max_column #最大列

#按行读取
for i in range(minrow,maxrow+1):
	if i>1:
		stu = Students()
		for j in range(mincol,maxcol+1):
					cell=sheet.cell(i,j).value
					print(cell,end=" ")
					if(j==1):
						print(f'cell1:{cell}')
						stu.classNum=cell
					elif(j==2):
						stu.xm=cell
					elif(j==3):
						stu.bmh=cell
					elif(j==4):
						stu.zkzh=cell
					elif(j==5):
						stu.sfzh=cell
		print(stu.bmh)

	else:
		print()